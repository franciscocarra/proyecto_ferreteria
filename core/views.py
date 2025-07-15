from django.shortcuts import render, redirect
from .models import *
from .forms import *
from django.contrib.auth.models import Group
from django.contrib.auth import authenticate, login, logout
import random
from .models import Producto
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import permission_required
from .models import Persona, TipoUsuario
import requests
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import bcrypt
from django.conf import settings
from django.urls import reverse
from transbank.webpay.webpay_plus.transaction import Transaction
from transbank.common.options import WebpayOptions
from transbank.common.integration_commerce_codes import IntegrationCommerceCodes
from transbank.common.integration_api_keys import IntegrationApiKeys
from transbank.common.integration_type import IntegrationType
from django.db.models import Prefetch, Sum
import time
from .models import Venta, DetalleVenta 
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime, timedelta
from io import BytesIO
from django.template.loader import get_template
from django.utils.crypto import get_random_string

API_URL = "http://127.0.0.1:8001/productos"

def home(request):
    return render(request, 'core/home.html')

def producto(request):
    try:
        # Hacemos una petición a la API para obtener todos los productos
        response = requests.get('http://127.0.0.1:8001/productos/')
        response.raise_for_status()  # Lanza un error si la petición falla
        productos = response.json()
    except requests.exceptions.RequestException:
        # Si la API no responde o hay un error, enviamos una lista vacía
        productos = []
    
    # Enviamos la lista de productos al template
    return render(request, 'core/productos.html', {'lista_productos': productos})



def contacto(request):
    return render(request, 'core/contacto.html')

@csrf_exempt
def iniciar_session(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('correo')
            password = data.get('contrasena')
        except (json.JSONDecodeError, AttributeError):
            return JsonResponse({'success': False, 'error': 'Datos inválidos.'}, status=400)

        user = authenticate(request, username=email, password=password)

        if user is not None:
            login(request, user)
            
            # Por defecto, redirigimos a home
            redirect_url = 'home'
            is_admin_login = False

            # 1. VERIFICACIÓN DE SUPERUSUARIO PRIMERO
            if user.is_superuser:
                redirect_url = 'administracion'
                is_admin_login = True
                # Nota: La lógica de cambio de contraseña no se ejecutará para superusers
                # a menos que también tengan un objeto Persona. Considera si esto es lo deseado.

            else:
                # 2. SI NO ES SUPERUSER, BUSCAMOS EL PERFIL
                try:
                    persona = Persona.objects.get(correo=user.email)
                    tipo = persona.tipo_persona.lower().strip()
                    
                    # ¡LÍNEA CLAVE PARA DEPURAR! Revisa la consola de Django.
                    print(f"Login attempt for {user.email}. Found user type: '{tipo}'")

                    if tipo == 'admin':
                        redirect_url = 'administracion'
                        is_admin_login = True
                    elif tipo == 'bodeguero':
                        redirect_url = 'bodeguero'
                    elif tipo == 'contador':
                        redirect_url = 'contador'
                    # Si el tipo no es ninguno de estos, se quedará en 'home'

                except Persona.DoesNotExist:
                    # Si no es superuser y no tiene perfil, no puede ser admin.
                    print(f"Login failed for {user.email}: No Persona object found.")
                    pass # redirect_url se mantiene como 'home'

            # --- LÓGICA PARA CAMBIAR CONTRASEÑA DE ADMIN ---
            # Se ejecuta solo si se confirmó que es un admin con perfil Persona
            if is_admin_login:
                try:
                    # Asegurarnos de que 'persona' existe para obtener el RUT
                    persona_for_pwd_change = Persona.objects.get(correo=user.email)
                    new_password = get_random_string(length=12)
                    
                    api_url = f"http://localhost:3000/personas/{persona_for_pwd_change.rut}"
                    api_key = 'Admin1234'
                    
                    response = requests.patch(
                        api_url,
                        headers={'Content-Type': 'application/json', 'x-api-key': api_key},
                        json={'contrasena': new_password}
                    )
                    
                    if response.ok:
                        request.session['new_admin_password'] = new_password
                        print(f"SUCCESS: Password for {user.email} updated via API.")
                    else:
                        print(f"API ERROR: Could not update password for {user.email}. Response: {response.text}")

                except Persona.DoesNotExist:
                     print(f"NOTICE: Cannot change password for superuser {user.email} without a Persona profile.")
                except Exception as e:
                    print(f"EXCEPTION during password update: {e}")
            # --- FIN DE LA LÓGICA ---

            return JsonResponse({'success': True, 'redirect_url_name': redirect_url})
        else:
            return JsonResponse({'success': False, 'error': 'Correo o contraseña incorrectos.'}, status=401)
    
    # El método GET no cambia
    if request.user.is_authenticated:
        return redirect('home')
    return render(request, 'core/iniciar_session.html')


def registro(request):
    # Ya no se necesita el formulario de Django aquí, porque la página usa
    # una API externa. Simplemente mostramos el HTML.
    return render(request, 'core/registro.html') 

    return render(request, 'core/registro.html', {'form': formulario})

def olvidar_contrasena(request):
    return render(request, 'core/olvidar_contrasena.html')

def pintura(request):
    return render(request, 'core/pintura.html')

def error_producto(request):
    return render(request, 'core/error_producto.html')

def herramientas(request):
    return render(request, 'core/herramientas.html')

def materiales(request):
    return render(request, 'core/materiales.html')

def carrito(request):
    cart = request.session.get('cart', {})
    cart_items = []
    total_cart_price = 0

    for sku, item_data in cart.items():
        # --- INICIO DE LA CORRECCIÓN ---
        # 1. Define la variable 'item_total' al principio del bucle.
        item_total = item_data['price'] * item_data['quantity']

        # 2. Ahora puedes usar 'item_total' para agregarlo a la lista de items.
        cart_items.append({
            'sku': sku,
            'name': item_data['name'],
            'price': item_data['price'],
            'quantity': item_data['quantity'],
            'image': item_data.get('image', 'core/img/placeholder.png'),
            'stock': item_data.get('stock', 0),
            'total': item_total  # Se usa aquí
        })
        
        # 3. Y también puedes usar 'item_total' para sumarlo al total general.
        total_cart_price += item_total # El error estaba aquí
        # --- FIN DE LA CORRECCIÓN ---
    
    context = {
        'cart_items': cart_items,
        'total_cart_price': total_cart_price
    }
    return render(request, 'core/carrito.html', context)


def add_to_cart(request, sku):
    try:
        response = requests.get(f'http://127.0.0.1:8001/productos/{sku}')
        response.raise_for_status()
        producto_data = response.json()
    except requests.exceptions.RequestException:
        messages.error(request, 'No se pudo obtener la información del producto.')
        return redirect('productos')

    if producto_data.get('stock', 0) > 0:
        cart = request.session.get('cart', {})
        sku_str = str(sku)

        if sku_str in cart:
            # ---> INICIO CAMBIO 1 <---
            # Solo aumenta la cantidad si es menor que el stock
            if cart[sku_str]['quantity'] < cart[sku_str]['stock']:
                cart[sku_str]['quantity'] += 1
            else:
                messages.warning(request, f"No puedes agregar más unidades de este producto. Stock máximo: {cart[sku_str]['stock']}")
            # ---> FIN CAMBIO 1 <---
        else:
            cart[sku_str] = {
                'name': producto_data['nombre_producto'],
                'price': float(producto_data['precio']),
                'quantity': 1,
                'image': 'core/img/productos/producto_1.webp',
                'stock': producto_data['stock'] # <-- AÑADE ESTA LÍNEA
            }
        
        request.session['cart'] = cart
        return redirect('carrito')
    else:
        messages.error(request, 'Este producto no tiene stock disponible.')
        return redirect('productos')

def remove_from_cart(request, sku):
    cart = request.session.get('cart', {})
    sku_str = str(sku)
    if sku_str in cart:
        del cart[sku_str]
        request.session['cart'] = cart
    return redirect('carrito')

def update_cart(request):
    if request.method == 'POST':
        sku = request.POST.get('sku')
        quantity = int(request.POST.get('quantity', 1))
        cart = request.session.get('cart', {})
        
        if sku in cart:
            stock_disponible = cart[sku].get('stock', 0)

            # ---> INICIO CAMBIO 2 <---
            # Limita la cantidad al stock disponible
            if quantity > stock_disponible:
                quantity = stock_disponible
                messages.warning(request, f"La cantidad se ha ajustado al stock máximo disponible: {stock_disponible}")

            if quantity > 0:
                cart[sku]['quantity'] = quantity
            else:
                del cart[sku]
            # ---> FIN CAMBIO 2 <---
        
        request.session['cart'] = cart
    return redirect('carrito')

def clear_cart(request):
    request.session['cart'] = {}
    return redirect('carrito')

@login_required
def iniciar_pago(request):
    print("\n--- 1. Entrando a la vista iniciar_pago ---")
    cart = request.session.get('cart', {})
    if not cart:
        print("--- ERROR: El carrito está vacío. Redirigiendo. ---")
        messages.error(request, "Tu carrito está vacío.")
        return redirect('carrito')

    total = sum(item['price'] * item['quantity'] for item in cart.values())
    buy_order = str(request.user.id) + "_" + str(int(time.time()))
    session_id = request.session.session_key
    amount = int(total)
    return_url = request.build_absolute_uri(reverse('confirmar_pago'))
    
    print(f"--- 2. Datos de la transacción listos ---")
    print(f"   - Orden de Compra: {buy_order}")
    print(f"   - Monto: {amount}")
    print(f"   - URL de Retorno: {return_url}")

    try:
        tx = Transaction(WebpayOptions(IntegrationCommerceCodes.WEBPAY_PLUS, IntegrationApiKeys.WEBPAY, IntegrationType.TEST))
        print("--- 3. Objeto de transacción de Transbank creado ---")
        
        response = tx.create(buy_order, session_id, amount, return_url)
        print("--- 4. Transacción creada en Transbank con éxito ---")
        print("   - Respuesta de Transbank:", response)
        
        # Leemos la respuesta como un diccionario
        url_transbank = response['url']
        token_transbank = response['token']
        
        print("   - URL de Redirección:", url_transbank)
        print("   - Token:", token_transbank)
        
        return redirect(url_transbank + '?token_ws=' + token_transbank)

    except Exception as e:
        print(f"--- 5. ERROR: La llamada a Transbank falló ---")
        print(f"   - Excepción: {e}")
        messages.error(request, f"Error al iniciar la transacción con Transbank.")
        return redirect('carrito')



def confirmar_pago(request):
    token = request.POST.get('token_ws') or request.GET.get('token_ws')
    
    if not token:
        messages.info(request, "Compra cancelada o token no encontrado.")
        print("--- DEBUG: Token no recibido o compra cancelada. Redirigiendo a carrito. ---")
        return redirect('carrito')

    print(f"--- DEBUG: Token recibido: {token} ---")

    try:
        # Intenta confirmar la transacción con Transbank
        print("--- DEBUG: Intentando confirmar la transacción con Transbank... ---")
        tx = Transaction(WebpayOptions(IntegrationCommerceCodes.WEBPAY_PLUS, IntegrationApiKeys.WEBPAY, IntegrationType.TEST))
        response = tx.commit(token)
        print(f"--- DEBUG: Respuesta de Transbank (commit): {response} ---")

        # --- INICIO DE LA CORRECCIÓN: Acceder a los elementos como diccionario ---
        # Verificar si la respuesta es un diccionario y acceder a sus claves
        if isinstance(response, dict) and response.get('status') == 'AUTHORIZED' and response.get('response_code') == 0:
            print("\n--- ¡PAGO APROBADO POR TRANSBANK! (Dentro de if 'AUTHORIZED') ---")
            
            try:
                print("--- 1. Intentando crear el objeto Venta... ---")
                # Asegúrate de que request.user esté autenticado y tenga un objeto Persona asociado si lo usas en la plantilla
                venta = Venta.objects.create(
                    usuario=request.user,
                    total=int(response.get('amount')), # Acceder como diccionario
                    orden_compra=response.get('buy_order'), # Acceder como diccionario
                    codigo_autorizacion=response.get('authorization_code'), # Acceder como diccionario
                    tipo_tarjeta=response.get('payment_type_code'), # Acceder como diccionario
                    # Usa .get() con un valor por defecto para evitar errores si 'card_number' no existe
                    ultimos_digitos_tarjeta=response.get('card_detail', {}).get('card_number', 'N/A') 
                )
                print(f"--- 2. Venta #{venta.id} creada exitosamente para usuario {request.user.username}. ---")

                cart = request.session.get('cart', {})
                print(f"--- 3. Carrito obtenido de la sesión. Items: {len(cart)} ---")

                if not cart:
                    print("--- ADVERTENCIA: Carrito vacío al momento de confirmar la venta. ---")
                    messages.warning(request, "El pago fue exitoso, pero el carrito estaba vacío. No se registraron productos.")

                for sku, item in cart.items():
                    print(f"    - Procesando producto SKU: {sku}")
                    try:
                        # Usar get_object_or_404 es más robusto para asegurar que el producto existe
                        producto = get_object_or_404(Producto, sku=sku) 
                        DetalleVenta.objects.create(
                            venta=venta,
                            producto=producto,
                            cantidad=item['quantity'],
                            precio_unitario=item['price']
                        )
                        print(f"    - Detalle para SKU: {sku} guardado.")
                    except Producto.DoesNotExist:
                        print(f"    --- ERROR: Producto con SKU '{sku}' no encontrado en la base de datos. ---")
                        messages.error(request, f"Producto '{sku}' no encontrado para registrar en la venta.")
                    except Exception as det_e:
                        print(f"    --- ERROR al guardar detalle para SKU '{sku}': {det_e} ---")
                        messages.error(request, f"Error al guardar detalle para un producto ({sku}).")


                print("--- 4. Todos los detalles de la venta procesados. ---")

            except Exception as e:
                print(f"--- ERROR CRÍTICO AL GUARDAR LA VENTA EN BASE DE DATOS: {e} ---")
                # Aunque el pago fue exitoso, no se pudo registrar la venta.
                messages.error(request, f"El pago fue exitoso, pero hubo un error al registrar tu compra. Por favor, contacta a soporte. Detalles: {e}")
                return redirect('home')


            # Vaciamos el carrito de la sesión
            if 'cart' in request.session:
                del request.session['cart']
                print("--- 5. Carrito vaciado de la sesión. ---")
            
            messages.success(request, "¡Tu compra ha sido realizada con éxito!")
            context = {'response': response}
            return render(request, 'core/pago_exitoso.html', context)
        else:
            # PAGO RECHAZADO por Transbank (response.status no es 'AUTHORIZED' o response_code no es 0)
            # También si response no es un diccionario o si el status no es 'AUTHORIZED'
            print(f"--- DEBUG: Pago rechazado por Transbank. Estado: {response.get('status', 'N/A')}, Código: {response.get('response_code', 'N/A')} ---")
            messages.error(request, "Tu pago fue rechazado. Inténtalo de nuevo.")
            context = {'response': response}
            return render(request, 'core/pago_fallido.html', context)
        # --- FIN DE LA CORRECCIÓN ---

    except Exception as e:
        # Este bloque captura cualquier error *antes* de que Transbank dé una respuesta de commit.
        # Por ejemplo, problemas de conexión, credenciales inválidas de Transbank, errores en la librería, etc.
        print(f"--- ERROR CRÍTICO EN LA COMUNICACIÓN CON TRANSBANK O EN LA LÓGICA PRINCIPAL: {e} ---")
        messages.error(request, f"Ocurrió un error inesperado al procesar tu pago. Por favor, contacta a soporte. Detalles: {e}")
        return redirect('carrito')

def users(request):
    return render(request, 'core/A_dmin/users.html')

@login_required
def administracion(request):
    # Lógica de seguridad para verificar si el usuario es admin
    tiene_permiso = False
    if request.user.is_superuser:
        tiene_permiso = True
    else:
        try:
            persona = Persona.objects.get(correo=request.user.email)
            if persona.tipo_persona.strip().lower() == 'admin':
                tiene_permiso = True
        except Persona.DoesNotExist:
            tiene_permiso = False
            
    if not tiene_permiso:
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('home')

    # Obtenemos la nueva contraseña de la sesión (y la eliminamos para que se muestre solo una vez)
    new_password = request.session.pop('new_admin_password', None)
    
    context = {
        'user': request.user,
        'new_password': new_password
    }
    
    return render(request, 'core/administración.html', context)


def bodeguero(request):
    return render(request, 'core/bodeguero/bode.html')

def inventario(request):
    try:
        # 1. Obtenemos los productos actuales de la API
        response = requests.get('http://127.0.0.1:8001/productos/')
        response.raise_for_status()
        productos_api = response.json()
        
        # Obtenemos una lista de los SKUs que vienen de la API
        skus_en_api = {prod_api['sku'] for prod_api in productos_api}

        # --- INICIO DE LA NUEVA LÓGICA DE ELIMINACIÓN ---
        
        # 2. Obtenemos los SKUs que tenemos en nuestra base de datos local
        skus_en_db = set(Producto.objects.values_list('sku', flat=True))
        
        # 3. Identificamos los SKUs que hay que borrar
        skus_para_borrar = skus_en_db - skus_en_api
        
        if skus_para_borrar:
            Producto.objects.filter(sku__in=skus_para_borrar).delete()
            print(f"Productos eliminados: {skus_para_borrar}")
            
        # --- FIN DE LA NUEVA LÓGICA DE ELIMINACIÓN ---

        # 4. Sincronizamos (actualizamos o creamos) el resto de los productos
        for prod_api in productos_api:
            Producto.objects.update_or_create(
                sku=prod_api['sku'],
                defaults={
                    'nombre_producto': prod_api['nombre_producto'],
                    'marca': prod_api['marca'],
                    'descripcion': prod_api.get('descripcion', ''),
                    'precio': prod_api['precio'],
                    'stock': prod_api['stock'],
                    'estado_producto': prod_api.get('estado_producto', 'Activo')
                }
            )

    except requests.exceptions.RequestException:
        print("Error: No se pudo conectar con la API de productos.")
    
    # El resto de la vista sigue igual, obteniendo los datos de la BD local
    query = request.GET.get('q')
    lista_productos = Producto.objects.all()

    if query:
        lista_productos = lista_productos.filter(nombre_producto__icontains=query)
    
    return render(request, 'core/bodeguero/inventario.html', {'lista_productos': lista_productos})

def ver_stock_sucursales(request, sku):
    try:
        producto = Producto.objects.get(sku=sku)
        stock_en_sucursales = StockPorSucursal.objects.filter(producto=producto)
    except Producto.DoesNotExist:
        producto = {'nombre_producto': 'Producto no encontrado'}
        stock_en_sucursales = []

    context = {
        'producto': producto,
        'stock_en_sucursales': stock_en_sucursales
    }
    
    return render(request, 'core/bodeguero/ver_stock_sucursales.html', context)

def registrar_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inventario')  # Cambia por la url de tu lista
    else:
        form = ProductoForm()

    return render(request, 'core/bodeguero/crear_p.html', {'form': form})

def eliminar_producto(request, sku):
    producto = get_object_or_404(Producto, sku=sku)
    producto.delete()
    return redirect('inventario')

def editar_producto_api(request, sku):
    """
    1) En GET: obtiene los datos del producto desde la API y los muestra en editar_p.html.
    2) En POST: toma los campos del formulario, arma los query params (incluyendo sku)
       y hace PUT a http://127.0.0.1:8001/productos?sku=...&nombre_producto=... etc.
    """

    # URL base de tu API (sin el SKU en la ruta)
    api_base = 'http://127.0.0.1:8001/productos'

    if request.method == 'POST':
        # Recolectamos todos los campos del formulario
        nombre = request.POST.get('nombre_producto')
        marca = request.POST.get('marca')
        descripcion = request.POST.get('descripcion')
        precio = request.POST.get('precio')
        stock = request.POST.get('stock')
        estado = request.POST.get('estado_producto')
        imagen_url = request.POST.get('imagen_url')

        # Aquí armamos el dict que irá en params (query string)
        data = {
            'sku': sku,
            'nombre_producto': nombre,
            'marca': marca,
            'descripcion': descripcion,
            'precio': precio,
            'stock': stock,
            'estado_producto': estado,
            'imagen_url' : imagen_url,
        }

        # Hacemos PUT enviando todo como query params
        try:
            respuesta = requests.put(f"{api_base}/{sku}", params=data)
        except requests.RequestException as e:
            return HttpResponse(f"Error de conexión con la API: {e}")

        if respuesta.status_code in (200, 204):
            # Actualización exitosa, volvemos al inventario
            return redirect('inventario')
        else:
            # API devolvió error, mostramos el texto que venga
            return render(request, 'core/bodeguero/editar_p.html', {
                'producto': data,  # para rellenar el formulario con lo que intentó guardarse
                'error': f"Error al actualizar desde la API: {respuesta.text}"
            })

    else:  # GET
        # Obtiene el producto desde la API para llenar el formulario
        try:
            respuesta = requests.get(f"{api_base}/{sku}")
        except requests.RequestException:
            return HttpResponse("Error de conexión al obtener datos del producto.")

        if respuesta.status_code == 200:
            producto = respuesta.json()
            return render(request, 'core/bodeguero/editar_p.html', {
                'producto': producto
            })
        else:
            return render(request, 'core/bodeguero/editar_p.html', {
                'error': "Producto no encontrado en la API."
            })


def editar_producto(request, sku):
    producto = get_object_or_404(Producto, sku=sku)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('inventario')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'core/bodeguero/editar_p.html', {'form': form, 'producto': producto})


def listar_empleados(request):
    tipos_empleado = TipoUsuario.objects.filter(descripcion__in=['vendedor', 'bodeguero'])
    empleados = Persona.objects.filter(tipo__in=tipos_empleado)
    return render(request, 'core/A_dmin/empleados/lista_e.html', {'lista_empleados': empleados})


@login_required
def lista_usuarios(request): # <-- Renombrada para mayor claridad
    # La lógica se mantiene: obtiene todos los registros de la tabla Persona que no son 'admin'
    usuarios = Persona.objects.exclude(tipo_persona__iexact='admin') # <-- Renombrada
    
    context = {
        'usuarios': usuarios # <-- Renombrada
    }
    return render(request, 'core/A_dmin/Cliente/lista_c.html', context)

@login_required
def agregar_usuario(request):
    # Esta vista no necesita lógica de formulario, solo muestra la página HTML.
    # La lógica real estará en el JavaScript de la plantilla.
    return render(request, 'core/A_dmin/Cliente/agregar_usuario.html')

def eliminar_usuario(request, id):
    # Obtener el objeto usuario con el id proporcionado
    usuario = get_object_or_404(Persona, id=id)

    # Eliminar el usuario
    usuario.delete()

    # Redirigir a la lista de usuarios
    return redirect('Clientes')

@login_required # 1. Solo usuarios logueados pueden acceder
def ventas(request):
    # 2. Verificamos si el usuario es un contador
    try:
        persona = Persona.objects.get(correo=request.user.email)
        if persona.tipo_persona.lower() != 'contador' and not request.user.is_superuser:
            messages.error(request, 'Acceso denegado. No tienes permisos de contador.')
            return redirect('home')
    except Persona.DoesNotExist:
        if not request.user.is_superuser:
            messages.error(request, 'Acceso denegado.')
            return redirect('home')

    # Obtenemos todas las ventas y traemos la información relacionada
    # para que la página cargue más rápido. Las más nuevas primero.
    # Se elimina 'usuario__persona' de prefetch_related porque no es una relación directa del modelo User.
    ventas_con_detalles = Venta.objects.prefetch_related(
        Prefetch('detalles', queryset=DetalleVenta.objects.select_related('producto'))
    ).order_by('-fecha')

    # --- INICIO DE LA CORRECCIÓN: Obtener y adjuntar el objeto Persona ---
    # Para cada Venta, obtenemos el objeto Persona asociado al usuario de la venta
    # y lo adjuntamos como 'persona_obj' para usarlo en la plantilla.
    for venta in ventas_con_detalles:
        if venta.usuario: # Aseguramos que haya un usuario vinculado a la venta
            try:
                # Buscamos el objeto Persona usando el correo electrónico del usuario
                venta.usuario.persona_obj = Persona.objects.get(correo=venta.usuario.email)
            except Persona.DoesNotExist:
                # Si no se encuentra la Persona, la establecemos como None
                venta.usuario.persona_obj = None 
        else:
            # Si no hay usuario, no hay Persona asociada
            venta.usuario.persona_obj = None 
    # --- FIN DE LA CORRECCIÓN ---

    # Pasamos el diccionario de contexto con 'ventas' a la plantilla
    return render(request, 'core/contador/ventas.html', {'ventas': ventas_con_detalles})

@login_required # 1. Solo usuarios logueados pueden acceder
def contador(request):
    # 2. Verificamos si el usuario es un contador
    try:
        persona = Persona.objects.get(correo=request.user.email)
        if persona.tipo_persona.lower() != 'contador' and not request.user.is_superuser:
            messages.error(request, 'Acceso denegado. No tienes permisos de contador.')
            return redirect('home')
    except Persona.DoesNotExist:
        if not request.user.is_superuser: # Permitir acceso a superusuarios por si acaso
            messages.error(request, 'Acceso denegado.')
            return redirect('home')
        
    # 3. Si tiene permiso, le mostramos la página del panel
    return render(request, 'core/contador/contador.html')

@login_required
def reporte_ventas(request):
    # Lógica de seguridad para verificar si el usuario es un contador o superusuario
    try:
        persona = Persona.objects.get(correo=request.user.email)
        if persona.tipo_persona.lower() != 'contador' and not request.user.is_superuser:
            messages.error(request, 'Acceso denegado. No tienes permisos para ver reportes.')
            return redirect('home')
    except Persona.DoesNotExist:
        if not request.user.is_superuser:
            messages.error(request, 'Acceso denegado.')
            return redirect('home')

    # --- Lógica para filtros de fecha ---
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    ventas_query = Venta.objects.all()

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            ventas_query = ventas_query.filter(fecha__gte=fecha_inicio)
        except ValueError:
            messages.error(request, "Formato de fecha de inicio inválido. Use AAAA-MM-DD.")

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            # Para incluir todo el día de la fecha final, sumamos un día y filtramos por menor que
            ventas_query = ventas_query.filter(fecha__lt=fecha_fin + timedelta(days=1))
        except ValueError:
            messages.error(request, "Formato de fecha de fin inválido. Use AAAA-MM-DD.")

    # Obtenemos las ventas con sus detalles y usuarios/personas relacionadas
    ventas_reporte = ventas_query.prefetch_related(
        Prefetch('detalles', queryset=DetalleVenta.objects.select_related('producto')),
    ).order_by('-fecha')

    # Adjuntar el objeto Persona a cada usuario de la venta (como hicimos en la vista 'ventas')
    for venta in ventas_reporte:
        if venta.usuario:
            try:
                venta.usuario.persona_obj = Persona.objects.get(correo=venta.usuario.email)
            except Persona.DoesNotExist:
                venta.usuario.persona_obj = None
        else:
            venta.usuario.persona_obj = None

    # Calcular el total general del reporte
    total_general_ventas = ventas_reporte.aggregate(Sum('total'))['total__sum'] or 0

    context = {
        'ventas': ventas_reporte,
        'total_general': total_general_ventas,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str,
    }
    return render(request, 'core/contador/reporte_ventas.html', context)


@login_required
def descargar_reporte_pdf(request):
    # Reutilizamos la lógica de obtención de datos de reporte_ventas
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    ventas_query = Venta.objects.all()

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            ventas_query = ventas_query.filter(fecha__gte=fecha_inicio)
        except ValueError:
            pass # Ignorar errores de formato para la descarga, ya se manejan en la vista principal

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            ventas_query = ventas_query.filter(fecha__lt=fecha_fin + timedelta(days=1))
        except ValueError:
            pass

    ventas_reporte = ventas_query.prefetch_related(
        Prefetch('detalles', queryset=DetalleVenta.objects.select_related('producto')),
    ).order_by('-fecha')

    for venta in ventas_reporte:
        if venta.usuario:
            try:
                venta.usuario.persona_obj = Persona.objects.get(correo=venta.usuario.email)
            except Persona.DoesNotExist:
                venta.usuario.persona_obj = None
        else:
            venta.usuario.persona_obj = None

    total_general_ventas = ventas_reporte.aggregate(Sum('total'))['total__sum'] or 0

    context = {
        'ventas': ventas_reporte,
        'total_general': total_general_ventas,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str,
        'request': request # Pasar el objeto request para {% static %} si se usa en la plantilla PDF
    }

    # Renderizar la plantilla HTML para PDF
    template_path = 'core/contador/reporte_ventas_pdf.html' # Usaremos una plantilla específica para PDF
    template = get_template(template_path)
    html = template.render(context)

    # Crear el PDF
    response = BytesIO()
    pdf = pisa.CreatePDF(
        html,
        dest=response,
        encoding='utf-8' # Asegura el soporte de caracteres especiales como 'ñ'
    )
    if not pdf.err:
        response = HttpResponse(response.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_ventas.pdf"'
        return response
    
    messages.error(request, "Error al generar el reporte PDF.")
    return redirect('reporte_ventas')


@login_required
def descargar_reporte_excel(request):
    # Reutilizamos la lógica de obtención de datos de reporte_ventas
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    ventas_query = Venta.objects.all()

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            ventas_query = ventas_query.filter(fecha__gte=fecha_inicio)
        except ValueError:
            pass

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            ventas_query = ventas_query.filter(fecha__lt=fecha_fin + timedelta(days=1))
        except ValueError:
            pass

    ventas_reporte = ventas_query.prefetch_related(
        Prefetch('detalles', queryset=DetalleVenta.objects.select_related('producto')),
    ).order_by('-fecha')

    for venta in ventas_reporte:
        if venta.usuario:
            try:
                venta.usuario.persona_obj = Persona.objects.get(correo=venta.usuario.email)
            except Persona.DoesNotExist:
                venta.usuario.persona_obj = None
        else:
            venta.usuario.persona_obj = None

    total_general_ventas = ventas_reporte.aggregate(Sum('total'))['total__sum'] or 0

    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # Estilos básicos
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal="center")

    # Cabeceras del reporte
    headers = [
        "ID Venta", "Fecha", "Cliente", "Email Cliente", "Orden Compra", 
        "Total Venta", "Método Pago", "Últimos Dígitos Tarjeta", "Código Autorización",
        "Productos Comprados"
    ]
    ws.append(headers)
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_aligned_text

    # Llenar datos
    for venta in ventas_reporte:
        cliente_nombre = "Usuario Desconocido"
        if venta.usuario and venta.usuario.persona_obj:
            cliente_nombre = f"{venta.usuario.persona_obj.nombre} {venta.usuario.persona_obj.appaterno}"
        elif venta.usuario:
            cliente_nombre = venta.usuario.username

        productos_str = ""
        for detalle in venta.detalles.all():
            productos_str += f"{detalle.producto.nombre_producto} (x{detalle.cantidad}) - ${detalle.precio_unitario} c/u; "
        
        row_data = [
            venta.id,
            venta.fecha.strftime("%d/%m/%Y %H:%M"),
            cliente_nombre,
            venta.usuario.email if venta.usuario else "N/A",
            venta.orden_compra,
            venta.total,
            venta.tipo_tarjeta,
            venta.ultimos_digitos_tarjeta,
            venta.codigo_autorizacion,
            productos_str.strip('; ') # Eliminar el último '; '
        ]
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Añadir total general
    ws.append([]) # Fila vacía
    ws.append(["", "", "", "", "Total General:", total_general_ventas, "", "", "", ""])
    ws[ws.max_row][5].font = bold_font # Celda del total general
    ws[ws.max_row][5].border = thin_border

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Guardar el libro en un buffer en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Crear la respuesta HTTP
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.xlsx"'
    return response

